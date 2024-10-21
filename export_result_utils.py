import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
from datetime import datetime
def read_test_results(file_path: str) -> pd.DataFrame:
    """
    đọc kết quả trong file test_results.text để clean và tạo pandas DataFrame
    gồm test_case_id và result (PASS, FAIL)
    Sample return:
    test_case_id                automation_result
0   API_GIVE_STAR_PRO.001              PASS
1   API_GIVE_STAR_PRO.002              PASS
2   API_GIVE_STAR_PRO.003              PASS
3   API_GIVE_STAR_PRO.004              PASS
4   API_GIVE_STAR_PRO.005              PASS
    """
    result_list = open(file_path, "r")
    rows = []
    for line in result_list:
        test_case_id = line.strip().split(":")[0][1:]
        test_case_result = line.strip().split(";")[1].strip()[1:-1]
        rows.append(
            {"test_case_id": test_case_id, "automation_result": test_case_result}
        )
    result_list.close()
    result_df = pd.DataFrame.from_records(rows)
    # print("our result: \n", result_df)
    return result_df
def read_excel_report(
    file_path: str, column_to_update: str, sheet_name: str
) -> pd.DataFrame:
    """
    đọc file excel report và trích xuất cột test_case_id và cột result cần update với automation test cases
    """
    id_testcase_column = 1  # số của cột chứa ID testcase (tính từ 0 từ trái sang) - thay đổi nếu format thay đổi
    skip_rows =  1   # skip n dòng đầu tạo ra format file - chỉ read data  --> cần thay đổi nếu format report thay đổi
    column_to_update =  int(column_to_update)  # số thứ tự của cột result cần update
    df = pd.read_excel(
        file_path,   
        skiprows=skip_rows,
        header=None,
        sheet_name=sheet_name,
    ).iloc[
        :, [id_testcase_column, column_to_update]    # lọc lấy cột id testcase và cột kết quả cần update
    ]  
    old_names = df.columns
    new_columns_name = ["test_case_id", "current_result"]
    df.rename(columns=dict(zip(old_names, new_columns_name)), inplace=True)
    return df
def update_report_and_save(
    input_text_file: str, input_excel_file: str, column_to_update: str, sheet_name: str
):
    """
    Update file test report với kết quả chạy automation test --> save vào file excel tạo mới
    """
    test_result: pd.DataFrame = read_test_results(input_text_file)
    excel_report_data: pd.DataFrame = read_excel_report(
        input_excel_file, column_to_update, sheet_name
    )
    updated_test_result: pd.DataFrame = excel_report_data.merge(
        test_result, how="left", on="test_case_id"
    )
    updated_test_result["final_update"] = updated_test_result.apply(
        lambda row: (
            row["automation_result"]
            if (pd.isnull(row["current_result"]))
            else row["current_result"]
        ),
        axis=1,
    )
    # print("updated_result size: \n", updated_test_result.shape)
    # print("updated_result: \n", updated_test_result)
    # chỉ chọn cột test result để update
    updated_result = updated_test_result[["final_update"]]
    # bắt đầu đọc excel workbook và viết kết quả
    wb = load_workbook(input_excel_file)
    ws = wb[sheet_name]
    data_rows = dataframe_to_rows(updated_result, index=False, header=False)
    skip_rows = 1   # skip những dòng tạo ra format file - chỉ read data  --> cần thay đổi nếu format report thay đổi
    row_start_to_write = skip_rows + 1
     
    for r_idx, row in enumerate(data_rows, row_start_to_write):
        for c_idx, value in enumerate(row, int(column_to_update) + 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    # tham số hóa file output vì không update trực tiếp vào file
    current_time = datetime.now().strftime("%Y_%m_%d_%H_%M_%S")
    output_path = (
        input_excel_file.split(".xlsx")[0] + f"_{current_time}" + "_updated.xlsx"
    )
    wb.save(output_path)
# test end to end flow - old format
# input_excel_file = "/home/nghiaht/workspace/rd-automation-framework/results_report/Portfolio_Testcase_sample_v0.1.xlsx"
# column_to_update = "8"
# input_text_file = "/home/nghiaht/workspace/rd-automation-framework/test_results.txt"
# sheet_name = "Chức năng 1"
# update_report_and_save(input_text_file, input_excel_file, column_to_update, sheet_name)
